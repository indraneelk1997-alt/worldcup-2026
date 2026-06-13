#!/usr/bin/env python3
"""
Parse the edited zone_battle_review.xlsx back into the nested zone_battle config,
VALIDATE it, and write a STAGING file (data/config/zone_battle.parsed.json).
Does NOT overwrite the live data/config/zone_battle.json.

Run:  uv run --with openpyxl python src/tools/parse_zone_battle_xlsx.py
"""
from __future__ import annotations
import json
from collections import OrderedDict
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[2]
XLSX = REPO / "zone_battle_review.xlsx"
OUT = REPO / "data" / "config" / "zone_battle.parsed.json"

VALID_ATTRS = {
    "finishing","shot_power","long_shots","penalties","heading_accuracy",
    "short_passing","long_passing","crossing","ball_control","vision",
    "def_awareness","standing_tackle","sliding_tackle","interceptions",
    "volleys","dribbling","curve","agility","balance","free_kick_accuracy",
    "positioning","composure","reactions","aggression",
    "acceleration","sprint_speed","jumping","stamina","strength",
}
VALID_FAMILIES = {"movement","finishing","passing","dribbling","defending","gk","set_piece"}

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["zones"]
rows = list(ws.iter_rows(min_row=2, values_only=True))
hdr = [c.value for c in ws[1]]
idx = {name: i for i, name in enumerate(hdr)}

warnings = []
# nested: zones[zone][context][stage] -> OrderedDict[duel_name] -> duel
zones: "OrderedDict" = OrderedDict()
for r in rows:
    if r is None or r[idx["zone"]] in (None, ""):
        continue
    zone = str(r[idx["zone"]]).strip()
    ctx = str(r[idx["context"]]).strip()
    stage = str(r[idx["stage"]]).strip()
    duel = str(r[idx["duel"]]).strip()
    side = str(r[idx["side"]]).strip()
    attr = str(r[idx["attribute"]]).strip()
    try:
        dw = float(r[idx["duel_w"]]); aw = float(r[idx["attr_w"]])
    except (TypeError, ValueError):
        warnings.append(f"{zone}/{ctx}/{stage}/{duel}: non-numeric weight -> skipped row")
        continue
    boosts_raw = r[idx["boost_families"]]
    boosts = [b.strip() for b in str(boosts_raw).split(";") if b and b.strip()] if boosts_raw else []

    if attr not in VALID_ATTRS:
        warnings.append(f"{zone}/{ctx}/{stage}/{duel}/{side}: UNKNOWN attribute {attr!r}")
    for b in boosts:
        if b not in VALID_FAMILIES:
            warnings.append(f"{zone}/{ctx}/{stage}/{duel}/{side}: UNKNOWN family {b!r}")

    z = zones.setdefault(zone, OrderedDict())
    c = z.setdefault(ctx, OrderedDict())
    s = c.setdefault(stage, OrderedDict())
    d = s.setdefault(duel, {"name": duel, "w": dw,
                            "att": {}, "att_boost": [], "def": {}, "def_boost": []})
    d["w"] = dw
    d[side][attr] = aw
    bkey = "att_boost" if side == "att" else "def_boost"
    if boosts and not d[bkey]:
        d[bkey] = boosts

# rebuild as plain dict with stages -> lists; check duel-weight sums
zones_out = OrderedDict()
for zone, ctxs in zones.items():
    zones_out[zone] = OrderedDict()
    for ctx, stages in ctxs.items():
        zones_out[zone][ctx] = OrderedDict()
        for stage, duels in stages.items():
            lst = list(duels.values())
            zones_out[zone][ctx][stage] = lst
            ssum = round(sum(x["w"] for x in lst), 3)
            if abs(ssum - 1.0) > 0.001:
                warnings.append(f"{zone}/{ctx}/{stage}: duel weights sum to {ssum} (not 1.0)")

# params
wp = wb["params"]
params = {r[0].value: r[1].value for r in wp.iter_rows(min_row=2)}
cfg = OrderedDict()
cfg["model_version"] = "zone_battle_v1"
cfg["_doc"] = "Parsed from zone_battle_review.xlsx (staging). See docs/item7_zone_battle.md."
cfg["approach_gate"] = float(params.get("approach_gate", 0.5))
cfg["family_mult"] = {"base": float(params.get("family_mult_base", 1.05)),
                      "plus": float(params.get("family_mult_plus", 1.10))}
cfg["zones"] = zones_out

OUT.write_text(json.dumps(cfg, indent=2))
print(f"wrote {OUT}")
print(f"zones={len(zones_out)}  duels={sum(len(s) for z in zones_out.values() for c in z.values() for s in c.values())}")
print(f"approach_gate={cfg['approach_gate']}  family_mult={cfg['family_mult']}")
if warnings:
    print(f"\n{len(warnings)} WARNING(S):")
    for w in warnings:
        print("  -", w)
else:
    print("\nno warnings - clean parse")
