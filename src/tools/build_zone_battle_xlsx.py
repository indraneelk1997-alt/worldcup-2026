#!/usr/bin/env python3
"""
Build an EDITABLE review spreadsheet for the item-7 zone-battle weights.

Flattens the 9 zones x 4 profiles (2 contexts x 2 stages x duels x 2 sides x
attributes) into one long, human-editable table. central_L1 + wing_L1 are read
LIVE from data/config/zone_battle.json (locked='Y'); the other 7 are first-draft
dicts below (locked='N'). Edit any cell, re-upload, and we parse it back to JSON.

Run:  uv run --with openpyxl python src/tools/build_zone_battle_xlsx.py
Out:  zone_battle_review.xlsx  (repo root)
"""
from __future__ import annotations
import json
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
CFG = json.loads((REPO / "data" / "config" / "zone_battle.json").read_text())
LOCKED = CFG["zones"]                       # central_L1, wing_L1 (authored + tuned)

# --- first-draft dicts for the other 7 zones (same schema as zone_battle.json) ---
def duel(name, w, att, ab, df, db):
    return {"name": name, "w": w, "att": att, "att_boost": ab, "def": df, "def_boost": db}

DRAFTS = {
  "central_L2": {
    "attack_vs_defense": {
      "approach": [
        duel("between_lines", 0.50, {"positioning":0.6,"reactions":0.4}, [], {"def_awareness":0.6,"reactions":0.4}, ["defending"]),
        duel("turn_shield",   0.50, {"ball_control":0.5,"balance":0.5}, ["dribbling"], {"strength":0.5,"aggression":0.5}, ["defending"]),
      ],
      "main": [
        duel("shoot_edge",   0.60, {"long_shots":0.6,"composure":0.4}, ["finishing"], {"standing_tackle":0.6,"def_awareness":0.4}, ["defending"]),
        duel("through_ball", 0.40, {"vision":0.6,"short_passing":0.4}, ["passing"], {"interceptions":0.6,"positioning":0.4}, ["defending"]),
      ],
    },
    "buildup_vs_pressure": {
      "approach": [
        duel("receive_under_press", 0.50, {"composure":0.5,"ball_control":0.5}, ["passing"], {"sprint_speed":0.5,"acceleration":0.5}, ["movement"]),
        duel("find_angle",          0.50, {"vision":0.6,"composure":0.4}, ["passing"], {"interceptions":1.0}, ["defending","movement"]),
      ],
      "main": [
        duel("progress_pass", 1.00, {"short_passing":0.6,"long_passing":0.4}, ["passing"], {"interceptions":0.5,"standing_tackle":0.5}, ["defending"]),
      ],
    },
  },
  "central_L3": {
    "attack_vs_defense": {
      "approach": [
        duel("find_space",   0.50, {"positioning":0.5,"vision":0.5}, ["passing"], {"def_awareness":0.6,"reactions":0.4}, ["defending"]),
        duel("receive_turn", 0.50, {"ball_control":0.5,"composure":0.5}, ["passing"], {"aggression":0.5,"strength":0.5}, ["defending"]),
      ],
      "main": [
        duel("progress_pass", 0.70, {"short_passing":0.4,"long_passing":0.3,"vision":0.3}, ["passing"], {"interceptions":0.6,"standing_tackle":0.4}, ["defending"]),
        duel("carry_through", 0.30, {"dribbling":0.6,"acceleration":0.4}, ["dribbling","movement"], {"standing_tackle":0.5,"def_awareness":0.5}, ["defending"]),
      ],
    },
    "buildup_vs_pressure": {
      "approach": [
        duel("receive_under_press", 0.50, {"composure":0.5,"ball_control":0.5}, ["passing"], {"sprint_speed":0.5,"acceleration":0.5}, ["movement"]),
        duel("scan_angle",          0.50, {"vision":0.6,"composure":0.4}, ["passing"], {"interceptions":1.0}, ["defending","movement"]),
      ],
      "main": [
        duel("progress_pass", 1.00, {"short_passing":0.6,"long_passing":0.4}, ["passing"], {"interceptions":0.5,"standing_tackle":0.5}, ["defending"]),
      ],
    },
  },
  "halfspace_L1": {
    "attack_vs_defense": {
      "approach": [
        duel("peel_off", 0.50, {"positioning":0.6,"reactions":0.4}, [], {"def_awareness":0.6,"reactions":0.4}, ["defending"]),
        duel("cut_in",   0.50, {"agility":0.4,"balance":0.3,"dribbling":0.3}, ["dribbling","movement"], {"standing_tackle":0.5,"strength":0.5}, ["defending"]),
      ],
      "main": [
        duel("finish_curl", 0.60, {"finishing":0.5,"curve":0.3,"composure":0.2}, ["finishing"], {"standing_tackle":0.6,"def_awareness":0.4}, ["defending"]),
        duel("cutback",     0.40, {"vision":0.6,"short_passing":0.4}, ["passing"], {"interceptions":0.6,"positioning":0.4}, ["defending"]),
      ],
    },
    "buildup_vs_pressure": {
      "approach": [
        duel("receive_under_press", 0.50, {"composure":0.5,"ball_control":0.5}, ["passing"], {"sprint_speed":0.5,"acceleration":0.5}, ["movement"]),
        duel("lane_space",          0.50, {"vision":0.6,"composure":0.4}, ["passing"], {"interceptions":1.0}, ["defending","movement"]),
      ],
      "main": [
        duel("play_out", 1.00, {"short_passing":0.6,"long_passing":0.4}, ["passing"], {"interceptions":0.5,"standing_tackle":0.5}, ["defending"]),
      ],
    },
  },
  "halfspace_L2": {
    "attack_vs_defense": {
      "approach": [
        duel("find_pocket",  0.50, {"positioning":0.5,"vision":0.5}, ["passing"], {"def_awareness":0.6,"reactions":0.4}, ["defending"]),
        duel("receive_turn", 0.50, {"ball_control":0.5,"agility":0.5}, ["dribbling","passing"], {"aggression":0.5,"strength":0.5}, ["defending"]),
      ],
      "main": [
        duel("key_pass",      0.60, {"vision":0.4,"short_passing":0.3,"long_passing":0.3}, ["passing"], {"interceptions":0.6,"positioning":0.4}, ["defending"]),
        duel("shoot_dribble", 0.40, {"long_shots":0.5,"dribbling":0.5}, ["finishing","dribbling"], {"standing_tackle":0.6,"def_awareness":0.4}, ["defending"]),
      ],
    },
    "buildup_vs_pressure": {
      "approach": [
        duel("receive_under_press", 0.50, {"composure":0.5,"ball_control":0.5}, ["passing"], {"sprint_speed":0.5,"acceleration":0.5}, ["movement"]),
        duel("find_angle",          0.50, {"vision":0.6,"composure":0.4}, ["passing"], {"interceptions":1.0}, ["defending","movement"]),
      ],
      "main": [
        duel("progress_pass", 1.00, {"short_passing":0.6,"long_passing":0.4}, ["passing"], {"interceptions":0.5,"standing_tackle":0.5}, ["defending"]),
      ],
    },
  },
  "halfspace_L3": {
    "attack_vs_defense": {
      "approach": [
        duel("find_space",   0.50, {"positioning":0.5,"vision":0.5}, ["passing"], {"def_awareness":0.6,"reactions":0.4}, ["defending"]),
        duel("receive_turn", 0.50, {"ball_control":0.5,"composure":0.5}, ["passing"], {"aggression":0.5,"strength":0.5}, ["defending"]),
      ],
      "main": [
        duel("progress_pass", 0.70, {"short_passing":0.4,"long_passing":0.3,"vision":0.3}, ["passing"], {"interceptions":0.6,"standing_tackle":0.4}, ["defending"]),
        duel("carry_through", 0.30, {"dribbling":0.6,"acceleration":0.4}, ["dribbling","movement"], {"standing_tackle":0.5,"def_awareness":0.5}, ["defending"]),
      ],
    },
    "buildup_vs_pressure": {
      "approach": [
        duel("receive_under_press", 0.50, {"composure":0.5,"ball_control":0.5}, ["passing"], {"sprint_speed":0.5,"acceleration":0.5}, ["movement"]),
        duel("scan_angle",          0.50, {"vision":0.6,"composure":0.4}, ["passing"], {"interceptions":1.0}, ["defending","movement"]),
      ],
      "main": [
        duel("progress_pass", 1.00, {"short_passing":0.6,"long_passing":0.4}, ["passing"], {"interceptions":0.5,"standing_tackle":0.5}, ["defending"]),
      ],
    },
  },
  "wing_L2": {
    "attack_vs_defense": {
      "approach": [
        duel("take_on", 0.50, {"agility":0.5,"balance":0.3,"dribbling":0.2}, ["dribbling","movement"], {"standing_tackle":0.4,"sliding_tackle":0.3,"def_awareness":0.3}, ["defending"]),
        duel("pace",    0.50, {"acceleration":0.6,"sprint_speed":0.4}, ["movement"], {"sprint_speed":0.6,"def_awareness":0.4}, ["defending","movement"]),
      ],
      "main": [
        duel("early_cross", 0.60, {"crossing":0.6,"curve":0.4}, ["passing"], {"standing_tackle":0.4,"sliding_tackle":0.3,"interceptions":0.3}, ["defending"]),
        duel("combine",     0.40, {"vision":0.5,"short_passing":0.5}, ["passing"], {"positioning":0.5,"interceptions":0.5}, ["defending"]),
      ],
    },
    "buildup_vs_pressure": {
      "approach": [
        duel("receive_under_press", 0.50, {"composure":0.5,"ball_control":0.5}, ["passing"], {"acceleration":0.5,"sprint_speed":0.5}, ["movement"]),
        duel("shield_escape",       0.50, {"balance":0.5,"agility":0.5}, ["dribbling"], {"aggression":0.5,"strength":0.5}, ["defending"]),
      ],
      "main": [
        duel("outlet", 0.70, {"short_passing":0.6,"long_passing":0.4}, ["passing"], {"interceptions":0.5,"standing_tackle":0.5}, ["defending"]),
        duel("carry",  0.30, {"dribbling":0.6,"acceleration":0.4}, ["dribbling","movement"], {"standing_tackle":0.4,"sliding_tackle":0.3,"sprint_speed":0.3}, ["defending","movement"]),
      ],
    },
  },
  "wing_L3": {
    "attack_vs_defense": {
      "approach": [
        duel("pace",  0.50, {"acceleration":0.6,"sprint_speed":0.4}, ["movement"], {"sprint_speed":0.6,"def_awareness":0.4}, ["defending","movement"]),
        duel("carry", 0.50, {"dribbling":0.5,"ball_control":0.5}, ["dribbling"], {"standing_tackle":0.5,"aggression":0.5}, ["defending"]),
      ],
      "main": [
        duel("progress_deliver", 0.70, {"crossing":0.5,"short_passing":0.5}, ["passing"], {"interceptions":0.5,"standing_tackle":0.5}, ["defending"]),
        duel("carry_continue",   0.30, {"dribbling":0.6,"acceleration":0.4}, ["dribbling","movement"], {"standing_tackle":0.5,"sprint_speed":0.5}, ["defending","movement"]),
      ],
    },
    "buildup_vs_pressure": {
      "approach": [
        duel("receive_under_press", 0.50, {"composure":0.5,"ball_control":0.5}, ["passing"], {"acceleration":0.5,"sprint_speed":0.5}, ["movement"]),
        duel("shield_escape",       0.50, {"balance":0.5,"agility":0.5}, ["dribbling"], {"aggression":0.5,"strength":0.5}, ["defending"]),
      ],
      "main": [
        duel("outlet",    0.70, {"short_passing":0.6,"long_passing":0.4}, ["passing"], {"interceptions":0.5,"standing_tackle":0.5}, ["defending"]),
        duel("carry_out", 0.30, {"dribbling":0.6,"acceleration":0.4}, ["dribbling","movement"], {"standing_tackle":0.4,"sliding_tackle":0.3,"sprint_speed":0.3}, ["defending","movement"]),
      ],
    },
  },
}

ORDER = ["central_L1","central_L2","central_L3",
         "halfspace_L1","halfspace_L2","halfspace_L3",
         "wing_L1","wing_L2","wing_L3"]
LANE = {"central":"Central","halfspace":"Half-space","wing":"Wing"}

# --- flatten to long rows ---
rows = []
for z in ORDER:
    lane, level = z.split("_")
    locked = "Y" if z in LOCKED else "N"
    zdef = LOCKED.get(z) or DRAFTS[z]
    for context in ("attack_vs_defense", "buildup_vs_pressure"):
        for stage in ("approach", "main"):
            for d in zdef[context][stage]:
                for side, bkey in (("att", "att_boost"), ("def", "def_boost")):
                    for attr, aw in d[side].items():
                        rows.append([z, LANE[lane], level, locked, context, stage,
                                     d["name"], d["w"], side, attr, aw,
                                     ";".join(d.get(bkey, []))])

# --- write workbook ---
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "zones"
hdr = ["zone","lane","level","locked","context","stage","duel","duel_w",
       "side","attribute","attr_w","boost_families"]
ws.append(hdr)
for r in rows:
    ws.append(r)

bold = Font(bold=True)
for c in ws[1]:
    c.font = bold
ws.freeze_panes = "A2"
lock_fill = PatternFill("solid", fgColor="DDF0DD")   # locked rows green
draft_fill = PatternFill("solid", fgColor="FFF6D5")  # draft rows yellow
for i, r in enumerate(rows, start=2):
    fill = lock_fill if r[3] == "Y" else draft_fill
    for c in ws[i]:
        c.fill = fill
widths = [14,11,6,7,20,9,18,7,5,16,7,22]
for col, w in zip("ABCDEFGHIJKL", widths):
    ws.column_dimensions[col].width = w

# params sheet
wp = wb.create_sheet("params")
wp.append(["param","value"])
wp.append(["approach_gate", CFG["approach_gate"]])
wp.append(["family_mult_base", CFG["family_mult"]["base"]])
wp.append(["family_mult_plus", CFG["family_mult"]["plus"]])
for c in wp[1]:
    c.font = bold
wp.column_dimensions["A"].width = 18

# legend sheet
wl = wb.create_sheet("legend")
notes = [
  ["HOW TO EDIT", ""],
  ["", "Each row = one attribute on one side of one duel. Edit attr_w / duel_w / boost_families,"],
  ["", "add or delete rows, rename duels. Keep duel_w consistent within a (zone,context,stage)."],
  ["", "side: att = in-possession (attacker/builder); def = out-of-possession (defender/presser)."],
  ["", "boost_families = ;-separated EA PlayStyle families that multiply that side's attrs in the duel."],
  ["", "locked='Y' rows (green) are tuned (central_L1, wing_L1); 'N' (yellow) are first-draft."],
  ["", ""],
  ["S27 BUCKETS (each attribute belongs to exactly one)", ""],
  ["ATTACK", "finishing, shot_power, long_shots, penalties, heading_accuracy"],
  ["POSSESSION", "short_passing, long_passing, crossing, ball_control, vision"],
  ["DEFENSE", "def_awareness, standing_tackle, sliding_tackle, interceptions"],
  ["SKILLS", "volleys, dribbling, curve, agility, balance, free_kick_accuracy"],
  ["IQ", "positioning, composure, reactions, aggression"],
  ["PHYSICAL", "acceleration, sprint_speed, jumping, stamina, strength"],
  ["", ""],
  ["FAMILIES", "movement, finishing, passing, dribbling, defending  (gk / set_piece parked)"],
  ["RESOLUTION", "per-duel Bradley-Terry; stage = weighted mean; threat = main*(g+(1-g)*approach)"],
]
for row in notes:
    wl.append(row)
wl["A1"].font = bold
wl["A8"].font = bold
wl.column_dimensions["A"].width = 30
wl.column_dimensions["B"].width = 80

out = REPO / "zone_battle_review.xlsx"
wb.save(out)
print(f"wrote {out}  ({len(rows)} rows, {len(ORDER)} zones)")
